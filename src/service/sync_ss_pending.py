"""
26SS 미입고 스타일 + 입고예정일 추출 (스타일 기준 집계)
엑셀 ● 2026 02 05 시트에서 실입고일(AB)이 없는 스타일의 ETA(AA) 추출
품번 기준으로 SKU를 집계하여 스타일 단위 데이터 생성
"""

import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_PATH = PROJECT_ROOT / "src" / "input" / "26SS_DV_schedule.xlsx"
JSON_PATH = PROJECT_ROOT / "frontend" / "public" / "data" / "duvetica_26s_pending_arrival.json"
SHEET_NAME = "● 2026 02 05"
DATA_START = 10


def extract_pending():
    print(f"Reading: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb[SHEET_NAME]

    # SKU 단위로 수집
    sku_rows: list[dict] = []
    for row in range(DATA_START, ws.max_row + 1):
        style_no = ws.cell(row=row, column=6).value  # F: STYLE NO.
        if not style_no:
            continue

        arrival_yn = ws.cell(row=row, column=1).value  # A: 입고여부
        actual_dt = ws.cell(row=row, column=28).value   # AB: 실입고일
        eta = ws.cell(row=row, column=27).value          # AA: 입고 예정일

        if arrival_yn == "종결" or actual_dt:
            continue

        eta_str = eta.strftime("%Y-%m-%d") if isinstance(eta, datetime) else None
        sn = str(style_no)
        # 리오더 제거한 기본 품번
        base_style = sn.replace("-리오더", "").replace("-REORDER", "").replace("-re", "")

        sku_rows.append({
            "style_no_raw": sn,
            "style_no": base_style,
            "style_name": ws.cell(row=row, column=8).value or "",
            "color": (ws.cell(row=row, column=9).value or "").strip(),
            "pcs": ws.cell(row=row, column=10).value or 0,
            "supplier": ws.cell(row=row, column=11).value or "",
            "eta": eta_str,
            "category": ws.cell(row=row, column=31).value or "",
            "spot": ws.cell(row=row, column=5).value or "",
        })

    # 스타일(품번) 기준 집계
    style_map: dict[str, dict] = {}
    for r in sku_rows:
        sn = r["style_no"]
        if sn not in style_map:
            style_map[sn] = {
                "style_no": sn,
                "style_name": r["style_name"],
                "supplier": r["supplier"],
                "category": r["category"],
                "spot": r["spot"],
                "eta": r["eta"],
                "total_pcs": 0,
                "colors": [],
                "sku_count": 0,
            }
        s = style_map[sn]
        s["total_pcs"] += r["pcs"]
        s["sku_count"] += 1
        if r["color"] and r["color"] not in s["colors"]:
            s["colors"].append(r["color"])
        # ETA: 가장 빠른 날짜 사용
        if r["eta"] and (not s["eta"] or r["eta"] < s["eta"]):
            s["eta"] = r["eta"]

    # KG 발주 데이터에 있는 스타일만 필터링 (시즌 이월 스타일 제외)
    kg_order_file = PROJECT_ROOT / "frontend" / "public" / "data" / "duvetica_26s_order_inbound.json"
    kg_style_nos: set[str] = set()
    if kg_order_file.exists():
        with open(kg_order_file, "r", encoding="utf-8") as f:
            kg_raw = json.load(f)
        kg_data = kg_raw["data"] if isinstance(kg_raw, dict) and "data" in kg_raw else kg_raw
        for r in kg_data:
            prdt_cd = r.get("PRDT_CD", "")
            if len(prdt_cd) > 4:
                kg_style_nos.add(prdt_cd[4:])  # V26SVDSK10461 -> VDSK10461

    if kg_style_nos:
        before = len(style_map)
        style_map = {sn: s for sn, s in style_map.items() if sn in kg_style_nos}
        filtered_out = before - len(style_map)
        if filtered_out:
            print(f"  KG 미존재 스타일 {filtered_out}건 제외 (시즌 이월 등)")

    styles = sorted(style_map.values(), key=lambda x: x["eta"] or "9999")

    output = {
        "meta": {
            "source": EXCEL_PATH.name,
            "synced_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "style_count": len(styles),
            "sku_count": sum(s["sku_count"] for s in styles),
        },
        "data": styles,
    }

    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False)

    print(f"Saved: {JSON_PATH.name} ({len(styles)} styles, {len(sku_rows)} SKUs)")
    return output


if __name__ == "__main__":
    extract_pending()
