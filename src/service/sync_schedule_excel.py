"""
생산스케줄 엑셀 → JSON 변환
OneDrive 동기화 폴더에서 읽어 frontend/public/data/에 저장

사용법:
  PYTHONPATH=. .venv/Scripts/python src/service/sync_schedule_excel.py
"""

import json
import sys
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl 설치 필요: uv pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DATA_DIR = PROJECT_ROOT / "frontend" / "public" / "data"

# OneDrive 경로
ONEDRIVE_DIR = Path(r"C:\Users\AD1305\OneDrive - F&F\F_ISO-듀베티카 공용파일 - 문서")

EXCEL_FILES = {
    "26S": ONEDRIVE_DIR / "■ 26SS_DV_생산스케줄 취합_260205.xlsx",
    "26F": ONEDRIVE_DIR / "■ 26FW_DV_생산스케줄 취합_260325 1.xlsx",
}

# 메인 시트명 패턴
SHEET_PATTERNS = ["● 2026", "● 2025"]


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def fmt_date(val) -> str | None:
    """날짜 값을 YYYY-MM-DD 문자열로 변환"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s in ("#VALUE!", "#N/A", "#REF!", "#CONNECT!"):
        return None
    return s[:10] if len(s) >= 10 else None


def find_main_sheet(wb) -> str | None:
    """메인 데이터 시트 찾기 (● 로 시작)"""
    for name in wb.sheetnames:
        if name.startswith("●"):
            return name
    return None


def parse_schedule(filepath: Path, season: str) -> list[dict]:
    """생산스케줄 엑셀을 파싱하여 레코드 리스트 반환"""
    # 파일 잠김 방지: 임시 복사
    import shutil
    import tempfile
    tmp = Path(tempfile.gettempdir()) / f"schedule_{season}.xlsx"
    shutil.copy2(filepath, tmp)

    wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
    sheet_name = find_main_sheet(wb)
    if not sheet_name:
        log(f"  [ERROR] 메인 시트(●) 없음: {filepath.name}")
        wb.close()
        return []

    ws = wb[sheet_name]
    log(f"  시트: {sheet_name}")

    # 시즌별 컬럼 매핑 (0-based index)
    # 26SS: ● 2026 02 05 시트
    COL_MAP_SS = {
        "status": 0, "spot": 4, "style_no": 5, "style_name": 7,
        "color": 8, "pcs": 9, "supplier": 10, "staff": 11,
        "supplier_eta": 12, "eta": 26, "actual_dt": 27,
        "remark": 33, "item_code": 29,
    }
    # 26FW: ● 26FW ORDER LIST 시트
    COL_MAP_FW = {
        "status": 0, "spot": 7, "style_no": 8, "style_name": 10,
        "color": 11, "pcs": 12, "supplier": 13, "staff": 14,
        "supplier_eta": 15, "eta": 33, "actual_dt": 34,
        "remark": 37, "item_code": 2,
    }

    is_fw = season.upper().endswith("F")
    col = COL_MAP_FW if is_fw else COL_MAP_SS

    def safe(row, idx):
        return row[idx] if len(row) > idx else None

    records = []
    for i, row in enumerate(ws.iter_rows(min_row=10, values_only=True)):
        if not row or len(row) < 20:
            continue

        status = str(safe(row, col["status"]) or "").strip()
        style_no = str(safe(row, col["style_no"]) or "").strip()
        if not style_no:
            continue

        # 캔슬 제외
        if status == "캔슬":
            continue

        style_name = str(safe(row, col["style_name"]) or "").strip()
        color = str(safe(row, col["color"]) or "").strip()
        pcs = safe(row, col["pcs"])
        supplier = str(safe(row, col["supplier"]) or "").strip()
        staff = str(safe(row, col["staff"]) or "").strip()
        supplier_eta = fmt_date(safe(row, col["supplier_eta"]))
        eta = fmt_date(safe(row, col["eta"]))
        actual_dt = fmt_date(safe(row, col["actual_dt"]))
        remark = str(safe(row, col["remark"]) or "").strip()
        spot = str(safe(row, col["spot"]) or "").strip()
        item_code = str(safe(row, col["item_code"]) or "").strip()

        # PCS 정리
        try:
            pcs_num = int(float(pcs)) if pcs else 0
        except (ValueError, TypeError):
            pcs_num = 0

        records.append({
            "season": season,
            "status": status,
            "style_no": style_no,
            "style_name": style_name,
            "color": color,
            "pcs": pcs_num,
            "supplier": supplier,
            "staff": staff,
            "spot": spot,
            "supplier_eta": supplier_eta,
            "eta": eta,
            "actual_dt": actual_dt,
            "remark": remark,
            "item_code": item_code,
        })

    wb.close()
    return records


def main():
    log("=" * 60)
    log("생산스케줄 엑셀 동기화 시작")
    log("=" * 60)

    all_records = []

    for season, filepath in EXCEL_FILES.items():
        log(f"\n[{season}] {filepath.name}")
        if not filepath.exists():
            log(f"  [SKIP] 파일 없음")
            continue

        records = parse_schedule(filepath, season)
        log(f"  파싱 완료: {len(records)}건")
        all_records.extend(records)

    if all_records:
        out = DATA_DIR / "duvetica_schedule.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(all_records, f, ensure_ascii=False)
        size_kb = out.stat().st_size / 1024
        log(f"\n저장: duvetica_schedule.json ({size_kb:.0f} KB, {len(all_records)} rows)")

    log("\n" + "=" * 60)
    log("동기화 완료!")
    log("=" * 60)


if __name__ == "__main__":
    main()
