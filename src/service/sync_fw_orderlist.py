"""
26FW ORDER LIST 엑셀 ↔ JSON 양방향 동기화
사용법:
  PYTHONPATH=. .venv/Scripts/python src/service/sync_fw_orderlist.py              # Excel → JSON
  PYTHONPATH=. .venv/Scripts/python src/service/sync_fw_orderlist.py --to-excel   # JSON → Excel
"""

import json
import sys
import argparse
from datetime import datetime
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_PATH = PROJECT_ROOT / "src" / "input" / "26FW_DV_schedule.xlsx"
JSON_PATH = PROJECT_ROOT / "frontend" / "public" / "data" / "duvetica_fw_orderlist.json"
SHEET_NAME = "● 26FW ORDER LIST"

HEADER_ROW_1 = 7
HEADER_ROW_2 = 8
DATA_START_ROW = 10
DATA_END_ROW = 323

# 컬럼 정의 (col_index, key, label, group, editable, type)
COLUMNS: list[dict] = [
    {"col": 1,  "key": "arrival_yn",      "label": "입고여부",       "group": "basic",    "editable": True,  "type": "text"},
    {"col": 2,  "key": "marketing",       "label": "마케팅",         "group": "basic",    "editable": True,  "type": "text"},
    {"col": 3,  "key": "item_type",       "label": "복종",           "group": "basic",    "editable": False, "type": "text"},
    {"col": 4,  "key": "category",        "label": "분류",           "group": "basic",    "editable": False, "type": "text"},
    {"col": 5,  "key": "gender",          "label": "성별",           "group": "basic",    "editable": False, "type": "text"},
    {"col": 6,  "key": "is_m",            "label": "M",              "group": "basic",    "editable": False, "type": "number"},
    {"col": 7,  "key": "sku_count",       "label": "SKU",            "group": "basic",    "editable": False, "type": "number"},
    {"col": 8,  "key": "spot",            "label": "SPOT",           "group": "basic",    "editable": False, "type": "text"},
    {"col": 9,  "key": "style_no",        "label": "STYLE NO.",      "group": "basic",    "editable": False, "type": "text"},
    {"col": 10, "key": "image",           "label": "IMAGE",          "group": "basic",    "editable": False, "type": "text"},
    {"col": 11, "key": "style_name",      "label": "STYLE NAME",     "group": "basic",    "editable": False, "type": "text"},
    {"col": 12, "key": "color",           "label": "COLOR",          "group": "basic",    "editable": False, "type": "text"},
    {"col": 13, "key": "pcs",             "label": "PCS",            "group": "basic",    "editable": False, "type": "number"},
    {"col": 14, "key": "supplier",        "label": "생산처",         "group": "info",     "editable": False, "type": "text"},
    {"col": 15, "key": "staff",           "label": "담당자",         "group": "info",     "editable": False, "type": "text"},
    {"col": 16, "key": "supplier_eta",    "label": "협력사 협의납기","group": "info",     "editable": True,  "type": "date_or_text"},
    # 원단완료
    {"col": 17, "key": "fabric_plan",     "label": "원단 예정일",    "group": "progress", "editable": True,  "type": "date"},
    {"col": 18, "key": "fabric_done",     "label": "원단 출고일",    "group": "progress", "editable": True,  "type": "date"},
    # 부자재
    {"col": 19, "key": "trim_plan",       "label": "부자재 예정일",  "group": "progress", "editable": True,  "type": "date"},
    {"col": 20, "key": "trim_done",       "label": "부자재 출고일",  "group": "progress", "editable": True,  "type": "date"},
    # QC
    {"col": 21, "key": "qc_plan",         "label": "QC 예정일",      "group": "progress", "editable": True,  "type": "date"},
    {"col": 22, "key": "qc_done",         "label": "QC 컨펌일",      "group": "progress", "editable": True,  "type": "date"},
    # PP
    {"col": 23, "key": "pp_plan",         "label": "PP 예정일",      "group": "progress", "editable": True,  "type": "date"},
    {"col": 24, "key": "pp_done",         "label": "PP 컨펌일",      "group": "progress", "editable": True,  "type": "date"},
    # 재단/편직
    {"col": 25, "key": "cutting_plan",    "label": "재단 계획",      "group": "progress", "editable": True,  "type": "date"},
    {"col": 26, "key": "cutting_done",    "label": "재단 실행",      "group": "progress", "editable": True,  "type": "date"},
    # 투입
    {"col": 27, "key": "putin_plan",      "label": "투입 계획",      "group": "progress", "editable": True,  "type": "date"},
    {"col": 28, "key": "putin_done",      "label": "투입 실행",      "group": "progress", "editable": True,  "type": "date"},
    # 생산완료
    {"col": 29, "key": "finish_plan",     "label": "생산완료 계획",  "group": "progress", "editable": True,  "type": "date"},
    {"col": 30, "key": "finish_done",     "label": "생산완료 실행",  "group": "progress", "editable": True,  "type": "date"},
    # 선적
    {"col": 31, "key": "ship_handover",   "label": "수납전달",       "group": "progress", "editable": True,  "type": "date"},
    {"col": 32, "key": "ship_plan",       "label": "선적계획",       "group": "progress", "editable": True,  "type": "date"},
    {"col": 33, "key": "ship_done",       "label": "선적실행",       "group": "progress", "editable": True,  "type": "date"},
    # 입고
    {"col": 34, "key": "arrival_plan",    "label": "입고 예정일",    "group": "progress", "editable": True,  "type": "date"},
    {"col": 35, "key": "arrival_done",    "label": "실입고일",       "group": "progress", "editable": True,  "type": "date"},
    # 비고/납기
    {"col": 36, "key": "remark",          "label": "협력사 비고",    "group": "remark",   "editable": True,  "type": "text"},
    {"col": 37, "key": "agreed_eta",      "label": "협의납기(MD)",   "group": "remark",   "editable": True,  "type": "date"},
    {"col": 38, "key": "md_history",      "label": "MD납기히스토리", "group": "remark",   "editable": True,  "type": "text"},
    {"col": 39, "key": "md_desired_eta",  "label": "MD희망납기",     "group": "md",       "editable": True,  "type": "date"},
    {"col": 40, "key": "pp_status",       "label": "PP",             "group": "md",       "editable": True,  "type": "text"},
    {"col": 41, "key": "pp_remark",       "label": "PP비고",         "group": "md",       "editable": True,  "type": "text"},
    {"col": 42, "key": "receipt",         "label": "수납",           "group": "md",       "editable": True,  "type": "text"},
    {"col": 43, "key": "full_wash",       "label": "완세탁",         "group": "md",       "editable": True,  "type": "text"},
    {"col": 44, "key": "inspection_rpt",  "label": "검사보고서",     "group": "md",       "editable": True,  "type": "text"},
    {"col": 45, "key": "arrival_qty",     "label": "입고수량",       "group": "md",       "editable": True,  "type": "number"},
    {"col": 46, "key": "retail_price",    "label": "판매가",         "group": "md",       "editable": True,  "type": "number"},
    {"col": 47, "key": "cost_price",      "label": "원가",           "group": "md",       "editable": True,  "type": "number"},
    {"col": 48, "key": "markup",          "label": "M/UP",           "group": "md",       "editable": False, "type": "number"},
    {"col": 49, "key": "po_date",         "label": "PO발행",         "group": "md",       "editable": True,  "type": "date"},
]

def cell_to_json(value, col_type: str):
    """셀 값을 JSON 직렬화 가능한 형태로 변환"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if col_type == "number":
        try:
            return float(value) if "." in str(value) else int(value)
        except (ValueError, TypeError):
            return None
    return str(value).strip()


def json_to_cell(value, col_type: str):
    """JSON 값을 셀에 쓸 값으로 변환"""
    if value is None or value == "":
        return None
    if col_type in ("date", "date_or_text"):
        if isinstance(value, str) and len(value) == 10:
            try:
                return datetime.strptime(value, "%Y-%m-%d")
            except ValueError:
                return value
        return value
    if col_type == "number":
        try:
            return float(value) if "." in str(value) else int(value)
        except (ValueError, TypeError):
            return value
    return str(value)


def excel_to_json():
    """Excel → JSON 변환"""
    print(f"Reading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb[SHEET_NAME]

    rows = []
    for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
        # 빈 행 건너뛰기 (style_no 없으면 스킵)
        style_no = ws.cell(row=row_idx, column=9).value
        if not style_no:
            continue

        record: dict = {"_row": row_idx}
        for col_def in COLUMNS:
            val = ws.cell(row=row_idx, column=col_def["col"]).value
            record[col_def["key"]] = cell_to_json(val, col_def["type"])
        rows.append(record)

    # 메타데이터 포함
    output = {
        "meta": {
            "source": str(EXCEL_PATH.name),
            "sheet": SHEET_NAME,
            "synced_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "row_count": len(rows),
        },
        "columns": COLUMNS,
        "data": rows,
    }

    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=None)

    size_kb = JSON_PATH.stat().st_size / 1024
    print(f"Saved: {JSON_PATH.name} ({size_kb:.0f} KB, {len(rows)} rows)")
    return output


def json_to_excel():
    """JSON → Excel 반영 (기존 엑셀에 덮어쓰기)"""
    print(f"Reading JSON: {JSON_PATH}")
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        payload = json.load(f)

    data = payload["data"]
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb[SHEET_NAME]

    updated = 0
    for record in data:
        row_idx = record.get("_row")
        if not row_idx:
            continue
        for col_def in COLUMNS:
            if not col_def["editable"]:
                continue
            key = col_def["key"]
            new_val = json_to_cell(record.get(key), col_def["type"])
            old_val = ws.cell(row=row_idx, column=col_def["col"]).value
            # 날짜 비교 정규화
            if isinstance(old_val, datetime) and isinstance(new_val, datetime):
                if old_val.date() == new_val.date():
                    continue
            elif old_val == new_val:
                continue
            ws.cell(row=row_idx, column=col_def["col"]).value = new_val
            updated += 1

    if updated > 0:
        wb.save(str(EXCEL_PATH))
        print(f"Excel updated: {updated} cells changed")
    else:
        print("No changes to write")

    return updated


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="26FW ORDER LIST sync")
    parser.add_argument("--to-excel", action="store_true", help="JSON → Excel 반영")
    args = parser.parse_args()

    if args.to_excel:
        json_to_excel()
    else:
        excel_to_json()
